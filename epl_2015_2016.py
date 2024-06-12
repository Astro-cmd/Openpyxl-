
from openpyxl import workbook, load_workbook
wb = load_workbook('epl_2015_2016.xlsx')
ws = wb.active
ws['H1'] = 'total goals'
#ws1 = wb.create_sheet('Manchester United')
start_row = 1
end_row = ws.max_row

total_goal = 0


for row in range(start_row, end_row +1):
    value1 = ws.cell(row=row, column=4).value
    value2 = ws.cell(row=row, column=5).value

    total_goal = value1 + value2

ws['I2'] = total_goal
                                             

wb.save('epl_2015_2016.xlsx')
# print(wb.sheetnames)
