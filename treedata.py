from openpyxl import workbook,  load_workbook
wb = load_workbook('Book1.xlsx')
ws= wb.active
# print(wb.sheetnames)
Trees = [["type",'leaf color','hieght'],['maple','red',549],['oak','green',783],['pine','green',1204],['Eucalyptus','green',786]]
# add the trees to the workbook
# for row in Trees:
    # ws.append(row)
ws['D1'] = 'total hieght'
from openpyxl.styles import Font  
ft = Font(bold=True)
for row in ws['A1:C1']:
    for cell in row:
        cell.font = ft
# specify the range of rows to process
column = 3
start_row = 2
end_row = ws.max_row 

toalHieght = 0
for row in range(start_row, end_row):
    value = ws.cell(row=row, column=column).value

    # ensure value is numeric
    if value is not None and isinstance(value, (int, float)):
        toalHieght += value

    # write the value in cell d2
    ws['D2'] = toalHieght

wb.save('Book1.xlsx')        